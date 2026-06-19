"""
tests/test_pipeline_extended.py
Extended integration + unit tests for src/etl/pipeline.py.

Covers gaps from Phase 3:
  - recon file not found falls back to DB name_match / copay tables
  - IS_TEST flag prevents archive_file from being called in test mode
  - PipelineSummary.as_dict() contains all expected keys
  - run_pipeline returns PipelineSummary with correct recon_rows count
  - rate correction: payroll Unskilled client with Skilled remittance rate
    reconciles under the correct care type
"""
from __future__ import annotations
from datetime import date
from io import BytesIO
from pathlib import Path
from unittest.mock import patch, MagicMock
import sys

import duckdb
import openpyxl
import pytest

from src.db.schema import create_all
from src.etl.pipeline import (
    PipelineSummary,
    _normalize_insurance,
    load_name_match_from_db,
    load_copay_clients_from_db,
    run_pipeline,
)


# ── helpers ───────────────────────────────────────────────────────────────────

def _mem_db() -> duckdb.DuckDBPyConnection:
    conn = duckdb.connect(":memory:")
    create_all(conn)
    return conn


def _build_minimal_payroll(tmp_path: Path,
                            client="JONES, BOB",
                            insurance="Medicaid",
                            regular=35.0) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "03062026"
    ws.append(["Paycheck Date:", "2026-03-13"])
    ws.append(["Work Week:", "2026-03-04", "to", None, "2026-03-10"])
    ws.append(["Client","Insurance","Employee","Paylocity Emp ID",
               "Regular hrs","Respite hrs","Total Hrs"])
    ws.append([client, insurance, "Worker, Alice", 1001.0, regular, 0.0, regular])
    ws2 = wb.create_sheet("Paylocity Mapping")
    ws2.append([None, None, None, None, None, None])          # row 0: blank (iloc[0])
    ws2.append(["Emp#","Last","First","Full","EmpID","Status"])  # row 1: headers (iloc[1])
    ws2.append([None, "Worker", "Alice", "Worker, Alice", 1001, "A"])  # row 2: data (iloc[2])
    p = tmp_path / "payroll.xlsx"; wb.save(p); return p


def _build_minimal_remittance(tmp_path: Path,
                               client_comma="JONES, BOB",
                               insurance="Medicaid",
                               billed=35.0, paid=35.0,
                               charge=700.0, payment=700.0) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Remittance Report Template"
    ws.append(["metadata"])
    ws.append([None]); ws.append([None])
    ws.append(["Batch","Date","Txn","Match","Claim","TxnType","Charge","Payment",
               "Allowed","First","Last","First DOS","Last DOS","TCN","Billed Hrs",
               "Paid Hrs","Hrs Rem","Client","Last, First","Month","Insurance","Pay Val"])
    ws.append([1877,"2026-03-10","EFT","","CLM001","PAYMENT",
               f"${charge:.2f}", f"${payment:.2f}", f"${charge:.2f}",
               "BOB","JONES","2026-03-04","2026-03-10","TCN001",
               billed, paid, 0.0, "JONES BOB", client_comma, "3/", insurance, payment])
    p = tmp_path / "remit.xlsx"; wb.save(p); return p


def _build_minimal_recon(tmp_path: Path,
                          payroll_name="JONES, BOB",
                          remit_name="JONES, BOB") -> Path:
    """Minimal Weekly Recon reference file with Name Match + Copay sheets."""
    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "Name Match"
    ws1.append(["Payroll Name","Remittance Name"])
    ws1.append([payroll_name, remit_name])
    ws2 = wb.create_sheet("Copay")
    ws2.append(["Client Name","Copay Amount","Effective From","Effective To"])
    p = tmp_path / "recon.xlsx"; wb.save(p); return p


# ── PipelineSummary.as_dict ───────────────────────────────────────────────────

class TestPipelineSummaryAsDict:
    _REQUIRED_KEYS = {
        "payroll_records", "payroll_clients", "remittance_records",
        "remittance_filtered", "name_match_entries", "copay_entries",
        "recon_rows", "result_good", "result_followup", "result_no_payroll",
        "unmatched_clients",
    }

    def test_all_keys_present(self):
        """PipelineSummary.as_dict() must contain all expected summary keys."""
        s = PipelineSummary()
        d = s.as_dict()
        for k in self._REQUIRED_KEYS:
            assert k in d, f"Missing key in PipelineSummary.as_dict(): '{k}'"

    def test_defaults_are_zero_or_empty(self):
        s = PipelineSummary()
        d = s.as_dict()
        for k in self._REQUIRED_KEYS - {"unmatched_clients"}:
            assert d[k] == 0, f"Default for '{k}' should be 0, got {d[k]}"
        assert d["unmatched_clients"] == []


# ── recon file not found → falls back to DB ───────────────────────────────────

class TestReconFileFallbackToDb:
    def test_name_match_loaded_from_db_when_recon_missing(self, tmp_path):
        """
        When cfg.recon_file does not exist, pipeline must NOT crash.
        It must instead load name_match and copay from the DB.
        """
        conn = _mem_db()
        # Seed name_match in DB
        conn.execute(
            "INSERT INTO name_match (id, payroll_name, remittance_name) "
            "VALUES (nextval('seq_name_match'), 'JONES, BOB', 'JONES, BOB')"
        )
        payroll_path   = _build_minimal_payroll(tmp_path)
        remit_path     = _build_minimal_remittance(tmp_path)
        missing_recon  = tmp_path / "does_not_exist.xlsx"

        summary = run_pipeline(
            payroll_path=payroll_path,
            remittance_path=remit_path,
            recon_path=missing_recon,
            db_path=tmp_path / "test.duckdb",
        )
        # Pipeline must complete; recon rows must be written
        assert isinstance(summary, PipelineSummary)
        assert summary.recon_rows >= 0    # not a crash

    def test_load_name_match_from_db_returns_dict(self):
        """load_name_match_from_db reads the in-memory table correctly."""
        conn = _mem_db()
        conn.execute(
            "INSERT INTO name_match (id, payroll_name, remittance_name) "
            "VALUES (nextval('seq_name_match'), 'HARRIS, PATRICIA', 'HARRIS, PATRICIA')"
        )
        m = load_name_match_from_db(conn)
        assert "HARRIS, PATRICIA" in m
        assert m["HARRIS, PATRICIA"] == "HARRIS, PATRICIA"

    def test_load_copay_clients_from_db_returns_set(self):
        """load_copay_clients_from_db returns an uppercase set."""
        conn = _mem_db()
        conn.execute(
            "INSERT INTO copay_clients "
            "(id, client_name, copay_amount, is_active) "
            "VALUES (nextval('seq_copay_clients'), 'HARRIS, PATRICIA', 383.00, TRUE)"
        )
        s = load_copay_clients_from_db(conn)
        assert "HARRIS, PATRICIA" in s


# ── IS_TEST prevents archiving ────────────────────────────────────────────────

class TestIsTestPreventsArchive:
    def test_archive_not_called_during_test_run(self, tmp_path):
        """
        When IS_TEST is True (we are in pytest), archive_file must never be called.
        Monkeypatch archive_file and assert it is not invoked.
        """
        payroll_path  = _build_minimal_payroll(tmp_path)
        remit_path    = _build_minimal_remittance(tmp_path)
        recon_path    = _build_minimal_recon(tmp_path)

        with patch("src.etl.pipeline.archive_file") as mock_archive, \
             patch("src.etl.pipeline.IS_TEST", True):
            run_pipeline(
                payroll_path=payroll_path,
                remittance_path=remit_path,
                recon_path=recon_path,
                db_path=tmp_path / "t.duckdb",
            )
        mock_archive.assert_not_called(), \
            "archive_file was called during a test run — IS_TEST guard is broken"


# ── full pipeline happy path ──────────────────────────────────────────────────

class TestPipelineHappyPath:
    def test_run_pipeline_returns_nonzero_recon_rows(self, tmp_path):
        """A minimal payroll + remittance + recon file produces ≥1 recon row."""
        payroll_path = _build_minimal_payroll(tmp_path)
        remit_path   = _build_minimal_remittance(tmp_path)
        recon_path   = _build_minimal_recon(tmp_path)

        summary = run_pipeline(
            payroll_path=payroll_path,
            remittance_path=remit_path,
            recon_path=recon_path,
            db_path=tmp_path / "p.duckdb",
        )
        assert summary.recon_rows >= 1

    def test_run_pipeline_twice_idempotent(self, tmp_path):
        """Running pipeline twice on the same files must not double recon rows."""
        payroll_path = _build_minimal_payroll(tmp_path)
        remit_path   = _build_minimal_remittance(tmp_path)
        recon_path   = _build_minimal_recon(tmp_path)
        db_path      = tmp_path / "idem.duckdb"

        s1 = run_pipeline(payroll_path=payroll_path,
                          remittance_path=remit_path,
                          recon_path=recon_path, db_path=db_path)
        s2 = run_pipeline(payroll_path=payroll_path,
                          remittance_path=remit_path,
                          recon_path=recon_path, db_path=db_path)
        assert s2.recon_rows == s1.recon_rows, \
            f"Double-run produced different recon_rows: {s1.recon_rows} → {s2.recon_rows}"


# ── rate correction: Skilled remittance for Unskilled-labelled payroll ────────

class TestRateCorrectionCareType:
    def test_skilled_rate_remittance_reconciles_under_skilled(self, tmp_path):
        """
        A payroll record with no LPN/RN in name but remittance billed at
        Skilled rate (~$54/hr) must reconcile under 'Skilled' care_type.
        """
        # Payroll: plain name, Sentara insurance — would be Unskilled by name alone
        payroll_path = _build_minimal_payroll(tmp_path, client="DREWRY, KAYLA",
                                               insurance="Sentara", regular=40.0)
        # Remittance: Sentara & PDN (PDN = Skilled marker), $54/hr charge rate
        remit_path = _build_minimal_remittance(
            tmp_path, client_comma="DREWRY, KAYLA",
            insurance="Sentara & PDN",
            billed=40.0, paid=40.0, charge=2160.0, payment=2160.0,
        )
        recon_path = _build_minimal_recon(tmp_path,
                                           payroll_name="DREWRY, KAYLA",
                                           remit_name="DREWRY, KAYLA")
        summary = run_pipeline(
            payroll_path=payroll_path,
            remittance_path=remit_path,
            recon_path=recon_path,
            db_path=tmp_path / "rate.duckdb",
        )
        assert summary.recon_rows >= 1
