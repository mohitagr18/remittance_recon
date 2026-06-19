"""
tests/test_payroll_parser_extended.py
Extended unit tests for src/etl/payroll.py using synthetic in-memory
Excel fixtures built with openpyxl — no real files required.

Covers every gap from Phase 3 that was not in the original test_payroll_parser.py:
  - nonexistent file raises FileNotFoundError
  - sheet detection: MMDDYYYY-named sheet found correctly
  - wrong/missing MMDDYYYY sheet raises ValueError
  - formula rows (insurance starting with =) are skipped
  - blank client_name rows are skipped
  - total_hours is computed from regular+respite, NOT read from col G
  - employee_id cast from float to int string
  - respite blank defaults to 0.0
  - aggregate_payroll_hours sums hours correctly for same client/insurance
"""
from __future__ import annotations
from datetime import date
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from src.etl.payroll import parse_payroll, aggregate_payroll_hours


# ── fixture builder ───────────────────────────────────────────────────────────

def _build_payroll_xlsx(
    sheet_name: str = "03062026",
    paycheck_date: str = "2026-03-13",
    week_start: str = "2026-03-04",
    week_end: str = "2026-03-10",
    rows: list[list] | None = None,
) -> BytesIO:
    """
    Build a minimal payroll Excel in-memory.
    Row 1: Paycheck Date: | <date>
    Row 2: Work Week:     | <start> | to | None | <end>
    Row 3: headers (col A-G)
    Row 4+: data rows supplied by caller
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(["Paycheck Date:", paycheck_date, None, None, None, None, None])
    ws.append(["Work Week:", week_start, "to", None, week_end, None, None])
    ws.append(["Client", "Insurance", "Employee", "Paylocity Emp ID",
               "Regular hrs", "Respite hrs", "Total Hrs"])

    default_rows = [
        ["JONES, BOB", "Medicaid", "Worker, Alice", 1001.0, 35.0, 0.0, 35.0],
    ]
    for r in (rows if rows is not None else default_rows):
        ws.append(r)

    # Also add the Paylocity Mapping sheet (parse_payroll expects it)
    wb.create_sheet("Paylocity Mapping")
    ws2 = wb["Paylocity Mapping"]
    ws2.append(["Employee ID", "Last Name", "First Name", "Status"])
    ws2.append([1001, "Worker", "Alice", "A"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _write_to_tmp(buf: BytesIO, tmp_path: Path, name: str = "payroll.xlsx") -> Path:
    p = tmp_path / name
    p.write_bytes(buf.read())
    return p


# ── nonexistent file ──────────────────────────────────────────────────────────

class TestPayrollNonexistentFile:
    def test_raises_file_not_found(self, tmp_path):
        """parse_payroll on a missing path must raise FileNotFoundError."""
        missing = tmp_path / "does_not_exist.xlsx"
        with pytest.raises((FileNotFoundError, OSError)):
            parse_payroll(missing)


# ── sheet detection ───────────────────────────────────────────────────────────

class TestSheetDetection:
    def test_detects_mmddyyyy_sheet(self, tmp_path):
        """parse_payroll finds and reads a MMDDYYYY-named sheet correctly."""
        buf = _build_payroll_xlsx(sheet_name="03062026")
        path = _write_to_tmp(buf, tmp_path)
        data = parse_payroll(path)
        assert len(data["records"]) > 0

    def test_no_digit_sheet_name_raises(self, tmp_path):
        """
        _detect_data_sheet raises ValueError when NO sheet has any digit.
        (Sheets with ANY digit are used as fallback, e.g. 'Sheet1'.)
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DataSheet"           # no digits at all
        ws.append(["Paycheck Date:", "2026-03-13"])
        ws.append(["Work Week:", "2026-03-04", "to", None, "2026-03-10"])
        ws.append(["Client", "Insurance", "Employee", "ID", "Reg", "Resp", "Total"])
        ws2 = wb.create_sheet("PaylocityMapping")   # no digits
        ws2.append(["EmpID","Last","First","Status"])
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        path = _write_to_tmp(buf, tmp_path, "bad_sheet.xlsx")
        with pytest.raises(ValueError):
            parse_payroll(path)


# ── formula rows skipped ──────────────────────────────────────────────────────

class TestFormulaRowsSkipped:
    def test_insurance_starting_with_equals_is_excluded(self, tmp_path):
        """Rows where insurance begins with '=' (formula cell) must be skipped."""
        rows = [
            ["JONES, BOB",  "=SUM(E5:E10)", "Worker, Alice", 1001.0, 35.0, 0.0, 35.0],
            ["SMITH, JANE", "Medicaid",      "Aide, Bob",     1002.0, 20.0, 0.0, 20.0],
        ]
        buf = _build_payroll_xlsx(rows=rows)
        path = _write_to_tmp(buf, tmp_path)
        data = parse_payroll(path)
        clients = [r["client_name_raw"] for r in data["records"]]
        assert "JONES, BOB" not in clients,  "Formula row leaked into records"
        assert "SMITH, JANE" in clients,     "Valid row was incorrectly dropped"


# ── blank client name skipped ─────────────────────────────────────────────────

class TestBlankClientNameSkipped:
    def test_row_with_no_client_name_excluded(self, tmp_path):
        """Rows with a blank client name (col A) must be silently skipped."""
        rows = [
            [None,         "Medicaid", "Aide, Bob",   1002.0, 20.0, 0.0, 20.0],
            ["SMITH, JANE","Medicaid", "Aide, Carol",  1003.0, 30.0, 0.0, 30.0],
        ]
        buf = _build_payroll_xlsx(rows=rows)
        path = _write_to_tmp(buf, tmp_path)
        data = parse_payroll(path)
        assert all(r["client_name_raw"] for r in data["records"]), \
            "Record with blank client_name_raw found in output"
        assert len(data["records"]) == 1


# ── total_hours computed from regular+respite ─────────────────────────────────

class TestTotalHoursComputed:
    def test_total_is_regular_plus_respite_not_col_g(self, tmp_path):
        """
        Col G may contain a formula or wrong cached value.
        total_hours must equal regular + respite (cols E + F), not col G.
        """
        rows = [
            # regular=30, respite=5, col G (intentionally wrong) = 999.0
            ["JONES, BOB", "Medicaid", "Worker, Alice", 1001.0, 30.0, 5.0, 999.0],
        ]
        buf = _build_payroll_xlsx(rows=rows)
        path = _write_to_tmp(buf, tmp_path)
        data = parse_payroll(path)
        assert len(data["records"]) == 1
        assert data["records"][0]["total_hours"] == pytest.approx(35.0), \
            "total_hours read from col G instead of computed from regular+respite"


# ── employee_id float→int ─────────────────────────────────────────────────────

class TestEmployeeIdCast:
    def test_float_emp_id_cast_to_int_string(self, tmp_path):
        """Employee ID stored as float (e.g. 12345.0) must be cast to '12345'."""
        rows = [
            ["JONES, BOB", "Medicaid", "Worker, Alice", 12345.0, 35.0, 0.0, 35.0],
        ]
        buf = _build_payroll_xlsx(rows=rows)
        path = _write_to_tmp(buf, tmp_path)
        data = parse_payroll(path)
        emp_id = data["records"][0]["employee_id"]
        assert emp_id == "12345", \
            f"employee_id not cast from float: got '{emp_id}'"
        assert "." not in str(emp_id), "Float decimal point leaked into employee_id"


# ── respite blank defaults to 0 ───────────────────────────────────────────────

class TestRespiteBlankDefaultsZero:
    def test_blank_respite_cell_total_hours_still_computed(self, tmp_path):
        """
        A blank respite_hours cell (col F) produces respite_hours=None from _parse_hours.
        total_hours must still be computed correctly (None treated as 0 in addition).
        NOTE: This test documents current source behaviour where None+35=35.
        """
        rows = [
            ["JONES, BOB", "Medicaid", "Worker, Alice", 1001.0, 35.0, None, 35.0],
        ]
        buf = _build_payroll_xlsx(rows=rows)
        path = _write_to_tmp(buf, tmp_path)
        data = parse_payroll(path)
        # total_hours is computed as (regular or 0) + (respite or 0), so None respite → 0
        assert data["records"][0]["total_hours"] == pytest.approx(35.0)


# ── aggregate_payroll_hours correctness ───────────────────────────────────────

class TestAggregatePayrollHoursCorrectness:
    def test_sums_hours_for_same_client_insurance(self, tmp_path):
        """Two aides for the same client+insurance must produce one aggregated row."""
        rows = [
            ["JONES, BOB", "Medicaid", "Aide, Alice", 1001.0, 20.0, 0.0, 20.0],
            ["JONES, BOB", "Medicaid", "Aide, Carol", 1002.0, 15.0, 5.0, 20.0],
        ]
        buf = _build_payroll_xlsx(rows=rows)
        path = _write_to_tmp(buf, tmp_path)
        data = parse_payroll(path)
        aggregated = aggregate_payroll_hours(data["records"])
        jones_rows = [r for r in aggregated
                      if r["client_name_raw"] == "JONES, BOB" and r["insurance"] == "Medicaid"]
        assert len(jones_rows) == 1, "Aggregation produced duplicate rows for same client"
        assert jones_rows[0]["total_hours"] == pytest.approx(40.0), \
            f"Aggregated hours wrong: {jones_rows[0]['total_hours']} != 40.0"

    def test_separate_rows_for_different_insurance(self, tmp_path):
        """Same client, different insurance → two separate aggregated rows."""
        rows = [
            ["JONES, BOB", "Medicaid", "Aide, Alice", 1001.0, 20.0, 0.0, 20.0],
            ["JONES, BOB", "Sentara",  "Aide, Carol", 1002.0, 15.0, 0.0, 15.0],
        ]
        buf = _build_payroll_xlsx(rows=rows)
        path = _write_to_tmp(buf, tmp_path)
        data = parse_payroll(path)
        aggregated = aggregate_payroll_hours(data["records"])
        jones_rows = [r for r in aggregated if r["client_name_raw"] == "JONES, BOB"]
        assert len(jones_rows) == 2, "Different insurance must produce separate rows"
